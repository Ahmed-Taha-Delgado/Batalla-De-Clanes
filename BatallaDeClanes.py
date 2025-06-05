import pygame 
from PIL import Image
from pythonds3.basic import Deque
import openpyxl
import random
import os

#Encuentra la direccion de la carpeta de imagenes
Base=os.path.dirname(__file__)
imagenesRuta=os.path.join(Base,"imagenes")

pygame.init()

#Resolución
anchoMenu=1280
altoMenu=720
anchoTarjeta=220
altoTarjeta=350
    
#Ventana
pygame.display.set_caption("Batalla De Clanes")
ventana=pygame.display.set_mode((anchoMenu,altoMenu))

#Colores
negro=(0,0,0)
blanco=(255,255,255)
verde=(75,175,75)
verdeOscuro=(47,87,44)
piedra=(100,100,100)
amarillo=(255,193,34)
gris=(205,205,205)
rojo=(216,36,41)
rojoIntenso=(139,0,0)
letras=(217,186,127)
cafeFondo=(90,63,22)
cafeBorde=(57,56,8)
colorBoton=(123,74,45)
colorBorde=(58,38,21)
colorTexto=(246,231,193)

#Fuentes
fuente1=pygame.font.SysFont("showcardgothic",160)
fuente1a=pygame.font.SysFont("showcardgothic",75)
fuente1b=pygame.font.SysFont("showcardgothic",50)
fuente1c=pygame.font.SysFont("showcardgothic",30)
fuente1d=pygame.font.SysFont("showcardgothic",18)
fuente2=pygame.font.SysFont("forte",60)
fuente3=pygame.font.SysFont("maturascriptcapitals",70)
fuente3a=pygame.font.SysFont("maturascriptcapitals",40)
fuente3b=pygame.font.SysFont("maturascriptcapitals",28)

#Direcciones de las imagenes usadas
imagenMenu=os.path.join(imagenesRuta,"fondoMenu.png")
imagen=Image.open(imagenMenu)
imagenInicio=os.path.join(imagenesRuta,"Inicio.png")
imagen2=Image.open(imagenInicio)
imagenGuerreros=os.path.join(imagenesRuta,"guerrero.png")
imagen3=Image.open(imagenGuerreros)
imagenArqueroR=os.path.join(imagenesRuta,"arqueroR.png")
imagenArqueroA=os.path.join(imagenesRuta,"arqueroA.png")
imagenEspadaA=os.path.join(imagenesRuta,"espadaA.png")
imagenEspadaR=os.path.join(imagenesRuta,"espadaR.png")
imagenGiganteR=os.path.join(imagenesRuta,"giganteR.png")
imagenGiganteA=os.path.join(imagenesRuta,"giganteA.png")
imagenCaballeroR=os.path.join(imagenesRuta,"caballeroR.png")
imagenCaballeroA=os.path.join(imagenesRuta,"caballeroA.png")
campoBatalla=os.path.join(imagenesRuta,"campo2.png")
imagenHistorial=os.path.join(imagenesRuta,"historial.png")
imagenFlechaArriba=os.path.join(imagenesRuta,"flechaarriba.png")
imagenFlechaAbajo=os.path.join(imagenesRuta,"flechaabajo.png")

rutaClan1=os.path.join(Base,"clan1.txt")
rutaClan2=os.path.join(Base,"clan2.txt")
rutaClan3=os.path.join(Base,"clan3.txt")
rutaAdmin=os.path.join(Base,"admin.txt")
rutaArco=os.path.join(Base,"arco.txt")
rutaCaballero=os.path.join(Base,"caballero.txt")
rutaEspada=os.path.join(Base,"espada.txt")
rutaGigante=os.path.join(Base,"gigante.txt")

rutaExcel=os.path.join(Base,"registroJuegos.xlsx")

#Archivos para funciones administrador
nivelAdmin=None

try:
    espadaDat=open(rutaEspada, "x")
    espadaDat.write("Nombre: Espadachin\n")
    espadaDat.write("Nivel: 1\n")
    espadaDat.write("Vida: 100\n")
    espadaDat.write("Danio: 20")
    espadaDat.close()
except FileExistsError:
    pass
try:
    arcoDat=open(rutaArco, "x")
    arcoDat.write("Nombre: Arquero\n")
    arcoDat.write("Nivel: 1\n")
    arcoDat.write("Vida: 70\n")
    arcoDat.write("Danio: 30")
    arcoDat.close()
except FileExistsError: 
    pass
try:
    giganteDat=open(rutaGigante, "x")
    giganteDat.write("Nombre: Gigante\n")
    giganteDat.write("Nivel: 1\n")
    giganteDat.write("Vida: 150\n")
    giganteDat.write("Danio: 15")
    giganteDat.close()
except FileExistsError:
    pass
try:
    caballeroDat=open(rutaCaballero, "x")
    caballeroDat.write("Nombre: Caballero\n")
    caballeroDat.write("Nivel: 1\n")
    caballeroDat.write("Vida: 120\n")
    caballeroDat.write("Danio: 18")
    caballeroDat.close()
except FileExistsError:
    pass
try:
    admi=open(rutaAdmin,"x")
    admi.write("normal\n")
    admi.write("facil")
    admi.close()
except FileExistsError:
    pass

#Crea o actualiza los diccionarios con los datos de las vidas, niveles y daños de las tropas de las funciones de administrador
def crearDiccionarios():
    global espadaBase,espadaRBase,arcoBase,arcoRBase,giganteBase,giganteRBase,caballeroBase,caballeroRBase
    global espada,espadaR,arco,arcoR,gigante,giganteR,caballero,caballeroR
    global guerreros, guerrerosDB, guerrerosRojos, enemigosDB

    espadaBase={}
    espadaRBase={}
    datos=open(rutaEspada,"r")
    for i,linea in enumerate(datos):
        dato,valor=linea.strip().split(":")
        dato=dato.strip()
        valor=valor.strip()
        if i==0:
            espadaBase[dato]=valor
            espadaRBase[dato]=valor
        else:
            valorMeter = int(valor)
            if valorMeter < 0:
                valorMeter=0
            espadaBase[dato]=valorMeter
            espadaRBase[dato]=valorMeter 
    datos.close()
    espadaBase["imagenes"]=imagenEspadaA
    espadaRBase["imagenes"]=imagenEspadaR

    arcoBase={}
    arcoRBase={}
    datos=open(rutaArco,"r")
    for i,linea in enumerate(datos):
        dato,valor=linea.strip().split(":")
        dato=dato.strip()
        valor=valor.strip()
        if i==0:
            arcoBase[dato]=valor 
            arcoRBase[dato]=valor
        else:
            valorMeter = int(valor)
            if valorMeter < 0:
                valorMeter=0
            arcoBase[dato]=valorMeter 
            arcoRBase[dato]=valorMeter

    datos.close()
    arcoBase["imagenes"]=imagenArqueroA
    arcoRBase["imagenes"]=imagenArqueroR

    giganteBase={}
    giganteRBase={}
    datos=open(rutaGigante, "r")
    for i,linea in enumerate(datos):
        dato,valor=linea.strip().split(":")
        dato=dato.strip()
        valor=valor.strip()
        if i==0:
            giganteBase[dato]=valor
            giganteRBase[dato]=valor  
        else:
            valorMeter = int(valor)
            if valorMeter < 0:
                valorMeter=0
            giganteBase[dato]=valorMeter
            giganteRBase[dato]=valorMeter
    datos.close()
    giganteBase["imagenes"]=imagenGiganteA 
    giganteRBase["imagenes"]=imagenGiganteR

    caballeroBase={}
    caballeroRBase={}
    datos=open(rutaCaballero,"r")
    for i,linea in enumerate(datos):
        dato,valor=linea.strip().split(":")
        dato=dato.strip()
        valor=valor.strip()
        if i==0:
            caballeroBase[dato]=valor  
            caballeroRBase[dato]=valor
        else:
            valorMeter = int(valor)
            if valorMeter < 0:
                valorMeter=0
            caballeroBase[dato]=valorMeter  
            caballeroRBase[dato]=valorMeter
    datos.close()
    caballeroBase["imagenes"]=imagenCaballeroA
    caballeroRBase["imagenes"]=imagenCaballeroR

    espada=espadaBase.copy()
    arco=arcoBase.copy()
    gigante=giganteBase.copy()
    caballero=caballeroBase.copy()

    espadaR=espadaRBase.copy()
    arcoR=arcoRBase.copy()
    giganteR=giganteRBase.copy()
    caballeroR=caballeroRBase.copy()

    guerreros=[espada,arco,gigante,caballero]
    guerrerosDB=[espadaBase,arcoBase,giganteBase,caballeroBase]
    guerrerosRojos=[espadaR,arcoR,giganteR,caballeroR]
    enemigosDB=[espadaRBase,arcoRBase,giganteRBase,caballeroRBase]

crearDiccionarios()

#variables globales
guerrerosElegidos=[]
vidasGuerreros=[]
vidasEnemigos=[]
ordenEnemigos=[]
enemigosRonda=[]
turnosVisibles=[]
imagenesTurnos=[]

tamanioCola=0
clanIngresado=0

colaActualizada=False
yaAtaco=False
elegirAtaque=False
vidaMaxima=False

textoContrasenia=""
datoAdmin=""
ultimaDireccion=""
textoIngresado=""
contrasenia="eda1"
sortear="Vida"
estado="menu"

#Cola de turnos
cola=Deque()

indice=0

#Función que se ejecuta cuando el administrador aumenta algun dato base
def aumentarDatosBase(atributo):
    #atributo="Vida" o "Danio"
    incremento=5
    #Lista con los archivos de los que obtendra la información
    archivos=[rutaEspada, rutaArco, rutaGigante, rutaCaballero]

    for archivo in archivos:
        leer=open(archivo,"r")
        lineas=leer.readlines()
        leer.close()

        #Lista donde se almacenara lo que se escribira en el archivo
        nuevoTexto=[]
        for linea in lineas:
            if linea.startswith(atributo+":"): #Verifica que la linea contenga el valor del atributo a modificar
                valorActual=int(linea.strip().split(":")[1]) #Linea strip en variable 1 porque almacena lo obtenido en una lista
                nuevoValor=valorActual+incremento
                nuevoTexto.append(atributo+": "+str(nuevoValor)+"\n")
            else:
                nuevoTexto.append(linea)

        #Se sobreescribe el archivo con el nuevo valor del atributo
        sobreEscribir=open(archivo,"w")
        sobreEscribir.writelines(nuevoTexto)
        sobreEscribir.close()
    crearDiccionarios()

#Función que se ejecuta cuando el administrador aum algun dato base
def disminuirDatosBase(atributo):
    #atributo="Vida" o "Danio"
    decremento=5
    #Lista con los archivos de los que obtendra la información
    archivos = [rutaEspada, rutaArco, rutaGigante, rutaCaballero] 

    for archivo in archivos:
        leer=open(archivo,"r")
        lineas=leer.readlines()
        leer.close()

        #Lista donde se almacenara lo que se escribira en el archivo
        nuevoTexto=[]
        for linea in lineas:
            if linea.startswith(atributo+":"): #Verifica que la linea contenga el valor del atributo a modificar
                valorActual=int(linea.strip().split(":")[1]) #Linea strip en variable 1 porque almacena lo obtenido en una lista
                nuevoValor=valorActual-decremento
                nuevoTexto.append(atributo+": "+str(nuevoValor)+"\n")
            else:
                nuevoTexto.append(linea)

        #Se sobreescribe el archivo con el nuevo valor del atributo
        sobreEscribir=open(archivo,"w")
        sobreEscribir.writelines(nuevoTexto)
        sobreEscribir.close()
    crearDiccionarios()

#Función que obtiene los datos para mostrarselos al admin cuando los quiera modificar
def obtenerDatos():
    archivos=[rutaEspada, rutaArco, rutaGigante, rutaCaballero]
    #Diccionario donde se  almacenaran los datos leidos
    datosPersonajes={}

    for archivo in archivos:
        cargar=open(archivo,"r")
        #Diccionario donde se almacenaran temporalmente los datos
        datos={}
        for linea in cargar:
            if ":" in linea: #Verifica que la linea del archivo contenga :
                clave,valor=linea.strip().split(":") #Parte la linea a partir del :, guardando la primera parte como clave y la segunda como valor
                datos[clave.strip()]=valor.strip() #Guarda en el diccionario datos la clave relacionada al valor
        nombre=datos.get("Nombre", archivo.replace(".txt", "")) #Hace nombre igual a la clave nombre de datos o en su defecto al nombre del archivo sin el txt
        datosPersonajes[nombre]=datos #Mete en datosPersonajes la variable nombre relacionada a datos

    return datosPersonajes

#Función que muestra los valores de los archivos de administrador sobre el daño y la vida 
def mostrarDatos(pantalla, fuente, datosPersonajes, tipoDeDato):
    #datosPersonajes es un diccionario
    #tipoDeDato="Vida" o "Danio"
    y=300  # Posición vertical inicial

    for nombre, datos in datosPersonajes.items(): #Obtiene nombre y datos del diccionario, nombre la clave, datos el valor
        valor=datos.get(tipoDeDato, "No disponible")
        texto=fuente.render(nombre+ " - "+tipoDeDato+": "+str(valor), True,blanco)
        pantalla.blit(texto,(900,y))
        y+=60  # Espaciado entre líneas

datosPersonajes=obtenerDatos()

#Función del menú de inicio
def menu():
    
    global boton1,botonCreditos,botonSalir,botonAdmin

    fondoMenu=pygame.image.load(imagenMenu)
    menuFondo=pygame.transform.scale(fondoMenu,(1280,720))
    fondoMenu=menuFondo
    
    #Objetos 
    titulo1=fuente1a.render("Batalla De Clanes",True,letras)
    tituloCreditos=fuente1b.render("Creditos",True,colorTexto)
    tituloSalir=fuente1b.render("Salir",True,colorTexto)
    jugar=fuente1b.render("jugar",True,colorTexto)
    textoAdmin=fuente1c.render("Administrador",True,colorTexto)
    
    fondo1=pygame.Rect(270,30,720,100)
    fondo1A=pygame.Rect(265,25,730,110)
    boton1=pygame.Rect((500,225,280,100))
    boton1A=pygame.Rect((495,220,290,110))
    botonCreditos=pygame.Rect(500,350,280,100)
    botonCreditosA=pygame.Rect(495,345,290,110)
    botonSalir=pygame.Rect(500,475,280,100)
    botonSalirA=pygame.Rect(495,470,290,110)
    botonAdmin=pygame.Rect(35,635,250,60)
    botonAdminA=pygame.Rect(30,630,260,70)

    #Objetos mostrados
    ventana.blit(fondoMenu,(0,0))

    pygame.draw.rect(ventana,cafeBorde,fondo1A,border_radius=10)
    pygame.draw.rect(ventana,cafeFondo,fondo1,border_radius=10)
    pygame.draw.rect(ventana,colorBorde,boton1A,border_radius=10)
    pygame.draw.rect(ventana,colorBorde,botonAdminA,border_radius=10)
    pygame.draw.rect(ventana,colorBorde,botonCreditosA,border_radius=10)
    pygame.draw.rect(ventana,colorBorde,botonSalirA,border_radius=10)
    pygame.draw.rect(ventana,colorBoton,boton1,border_radius=10)
    pygame.draw.rect(ventana,colorBoton,botonCreditos,border_radius=10)
    pygame.draw.rect(ventana,colorBoton,botonSalir,border_radius=10)
    pygame.draw.rect(ventana,colorBoton,botonAdmin,border_radius=10)

    ventana.blit(titulo1,(279,47))
    ventana.blit(jugar,(560,255))
    ventana.blit(tituloCreditos,(520,377))
    ventana.blit(tituloSalir,(570,502))
    ventana.blit(textoAdmin,(38,652))

#Funcion de ingresar la contraseña del administrador
def contraseñaAdmin():

    global regresar4
    global textoContrasenia

    #textos y botones
    fondoMenu=pygame.image.load(imagenMenu)
    menuFondo=pygame.transform.scale(fondoMenu,(1280,720))
    fondoMenu=menuFondo
    
    texto=fuente1b.render("Ingresa la contraseña de administrador:",True,letras)
    textoRegresar2=fuente1c.render("Regresar",True,colorTexto)
    textoMostrar=fuente1b.render(textoContrasenia, True, negro)

    ingresarR=pygame.Rect(30,85,1220,100)
    ingresarRA=pygame.Rect(25,80,1230,110)
    rect1=pygame.Rect(100,250,1080,160)
    rect2=pygame.Rect(110,260,1060,140)
    regresar4=pygame.Rect(990,620,200,70)
    regresar4A=pygame.Rect(985,615,210,80)

    #Mostrar los botones y textos
    ventana.blit(fondoMenu,(0,0))

    pygame.draw.rect(ventana,cafeBorde,ingresarRA,border_radius=10)
    pygame.draw.rect(ventana,cafeFondo,ingresarR,border_radius=10)
    pygame.draw.rect(ventana,piedra,rect1,border_radius=10)
    pygame.draw.rect(ventana,gris,rect2,border_radius=10)
    pygame.draw.rect(ventana,colorBorde,regresar4A,border_radius=10)
    pygame.draw.rect(ventana,colorBoton,regresar4,border_radius=10)
    
    ventana.blit(texto,(88,104))
    ventana.blit(textoRegresar2,(1015,643))
    ventana.blit(textoMostrar, (150, 310))

#Funcion que muestra la interfaz del administrador   
def admin():
    
    global regresarMenu
    global indicarNivel,indicarVida,indicarDanio,indicarDificultad
    global alto,medio,bajo
    global facil,dificil
    global masVida,menosVida,masDanio,menosDanio

    #Interfaz del menu de administrador
    fondoMenu=pygame.image.load(imagenMenu)
    menuFondo=pygame.transform.scale(fondoMenu,(1280,720))
    fondoMenu=menuFondo

    regresarMenu=pygame.Rect(990,620,200,80)
    regresarMenuA=pygame.Rect(985,615,210,90)
    indicar=pygame.Rect(40,40,1200,90)
    indicarA=pygame.Rect(35,35,1210,100)
    textoRegresar2=fuente1c.render("Regresar",True,colorTexto)
    textoIndicar=fuente1b.render("¿Qué valor modificarás de los guerreros?",True,letras)

    textoVida=fuente1b.render("Vida",True,colorTexto)
    textoDanio=fuente1b.render("Daño",True,colorTexto)
    textoEscalado=fuente1b.render("Escalado",True,colorTexto)
    textoDificultad=fuente1b.render("Dificultad",True,colorTexto)

    indicarVida=pygame.Rect(100,190,340,100)
    indicarNivelA=pygame.Rect(95,185,350,110)
    indicarDanio=pygame.Rect(100,310,340,100)
    indicarVidaA=pygame.Rect(95,305,350,110)
    indicarNivel=pygame.Rect(100,430,340,100)
    indicarDanioA=pygame.Rect(95,425,350,110)
    indicarDificultad=pygame.Rect(100,550,340,100)
    indicarDificultadA=pygame.Rect(95,545,350,110)

    bajo=pygame.Rect(570,430,140,100) 
    bajoA=pygame.Rect(565,425,150,110) 
    medio=pygame.Rect(750,430,140,100)
    medioA=pygame.Rect(745,425,150,110) 
    alto=pygame.Rect(930,430,140,100) 
    altoA=pygame.Rect(925,425,150,110)  

    facil=pygame.Rect(570,550,150,100)
    facilA=pygame.Rect(565,545,160,110)
    dificil=pygame.Rect(760,550,150,100)
    dificilA=pygame.Rect(755,545,160,110)
    
    masVida=pygame.Rect(570,190,100,100)
    masVidaA=pygame.Rect(565,185,110,110)
    menosVida=pygame.Rect(700,190,100,100)
    menosVidaA=pygame.Rect(695,185,110,110)

    masDanio=pygame.Rect(570,310,100,100)
    masDanioA=pygame.Rect(565,305,110,110)
    menosDanio=pygame.Rect(700,310,100,100)
    menosDanioA=pygame.Rect(695,305,110,110)

    textoMasV=fuente1a.render("-",True,negro)
    textoMenosV=fuente1a.render("+",True,negro)
    
    textoMasD=fuente1a.render("-",True,negro)
    textoMenosD=fuente1a.render("+",True,negro)

    textoBajo=fuente1c.render("bajo",True,negro)
    textoMedio=fuente1c.render("medio",True,negro)
    textoAlto=fuente1c.render("alto",True,negro)

    textoFacil=fuente1c.render("fácil",True,negro)
    textoDificil=fuente1c.render("difícil",True,negro)
    
    #Objetos mostrados del menu de administrador
    ventana.blit(fondoMenu,(0,0))
    pygame.draw.rect(ventana,colorBorde,regresarMenuA,border_radius=10)
    pygame.draw.rect(ventana,colorBoton,regresarMenu,border_radius=10)
    pygame.draw.rect(ventana,cafeBorde,indicarA,border_radius=10)
    pygame.draw.rect(ventana,cafeFondo,indicar,border_radius=10)

    pygame.draw.rect(ventana,colorBorde,indicarNivelA,border_radius=10)
    pygame.draw.rect(ventana,colorBorde,indicarDanioA,border_radius=10)
    pygame.draw.rect(ventana,colorBoton,indicarNivel,border_radius=10)
    pygame.draw.rect(ventana,colorBorde,indicarVidaA,border_radius=10)
    pygame.draw.rect(ventana,colorBoton,indicarDanio,border_radius=10)
    pygame.draw.rect(ventana,colorBoton,indicarVida,border_radius=10)
    pygame.draw.rect(ventana,colorBorde,indicarDificultadA,border_radius=10)
    pygame.draw.rect(ventana,colorBoton,indicarDificultad,border_radius=10)

    #Segun el boton seleecionado se muestran lo botones
    if datoAdmin=="escalado":
        pygame.draw.rect(ventana,colorBorde,bajoA,border_radius=10)
        pygame.draw.rect(ventana,verde,bajo,border_radius=10)
        pygame.draw.rect(ventana,colorBorde,medioA,border_radius=10)
        pygame.draw.rect(ventana,amarillo,medio,border_radius=10)
        pygame.draw.rect(ventana,colorBorde,altoA,border_radius=10)
        pygame.draw.rect(ventana,rojo,alto,border_radius=10)

        ventana.blit(textoBajo,(605,470))
        ventana.blit(textoMedio,(775,470))
        ventana.blit(textoAlto,(960,470))
    
    if datoAdmin=="dificultad":
        pygame.draw.rect(ventana,colorBorde,facilA,border_radius=10)
        pygame.draw.rect(ventana,verde,facil,border_radius=10)
        pygame.draw.rect(ventana,colorBorde,dificilA,border_radius=10)
        pygame.draw.rect(ventana,rojo,dificil,border_radius=10)

        ventana.blit(textoFacil,(605,587 ))
        ventana.blit(textoDificil,(785,587))

    if datoAdmin=="vida":
        pygame.draw.rect(ventana,colorBorde,masVidaA,border_radius=10)
        pygame.draw.rect(ventana,rojo,masVida,border_radius=10)
        pygame.draw.rect(ventana,colorBorde,menosVidaA,border_radius=10)
        pygame.draw.rect(ventana,verde,menosVida,border_radius=10)

        ventana.blit(textoMasV,(605,205))
        ventana.blit(textoMenosV,(730,205))

    if datoAdmin=="danio":
        pygame.draw.rect(ventana,colorBorde,masDanioA,border_radius=10)
        pygame.draw.rect(ventana,rojo,masDanio,border_radius=10)    
        pygame.draw.rect(ventana,colorBorde,menosDanioA,border_radius=10)
        pygame.draw.rect(ventana,verde,menosDanio,border_radius=10)

        ventana.blit(textoMasD,(605,325))
        ventana.blit(textoMenosD,(730,325))

    ventana.blit(textoRegresar2,(1015,643))
    ventana.blit(textoIndicar,(55,55))

    ventana.blit(textoVida,(212,215))
    ventana.blit(textoDanio,(210,325))
    ventana.blit(textoEscalado,(152,455))
    ventana.blit(textoDificultad,(130,575))

    if nivelAdmin!=None:
        mostrarDatos(ventana,fuente1c,datosPersonajes,nivelAdmin)

#Función que mejora a los enemigos segun la dificultad y la tropa mas alta del jugado¿   
def mejorarEnemigos(personajes,datosBase,archivo):
    masAlto=1
    datos=open(archivo,"r")
    for linea in datos: 
        if ":" not in linea:
            continue
        nombre,nivel=linea.strip().split(":")
        if masAlto<int(nivel.strip()): #Hacemos masAlto igual al nivel mas alto de las tropas del jugador
            masAlto=int(nivel.strip())
    datos.close()

    #Se abre el archivo de adminstrador para evaluar que escalado tendran los enemigos
    admi=open(rutaAdmin, "r")
    lineas = admi.readlines()
    if len(lineas) >= 2:
        escalado = lineas[1].strip()
    else:
        escalado="bajo"

    if escalado=="bajo":
        escaladoVida=0.15
        escaladoDanio=0.1
    elif escalado=="normal":
        escaladoVida=0.25
        escaladoDanio=0.2
    elif escalado=="alto":
        escaladoVida=0.35
        escaladoDanio=0.3
    
    print(escalado)
        
    #Para cada guerrero de la lista de enemigos
    for tropa, base in zip(personajes, datosBase): #Recorremos personajes y datosBase al mismo tiempo usando un indice para cada uno
        #Aumentamos los datos segun la base y la dificultad obtenido
        tropa["Danio"] = round(base["Danio"] * (1 + escaladoDanio * (masAlto - 1)), 2)
        tropa["Vida"] = round(base["Vida"] * (1 + escaladoVida * (masAlto - 1)), 2)
        tropa["Nivel"] = masAlto

#Función que aumenta el daño y vida de los personajes segun su nivel y loa valores base       
def mejorarPersonajes(personajes,datosBase,archivo):
    #Diccionario donde se almacenaran los niveles de cada tropa
    niveles={}
    datos=open(archivo, "r")
    for linea in datos:
        if ":" not in linea:
            continue
        nombre,nivel=linea.strip().split(":")
        niveles[nombre]=int(nivel.strip()) #Guarda en niveles el nombre de la tropa como clave de su nivel
    datos.close()
    # Aplicar mejoras a cada tropa
    for tropa,base in zip(personajes,datosBase): #Recorremos personajes y datosBase al mismo tiempo usando un indice para cada uno
        tipoDeTropa=tropa["Nombre"]
        if tipoDeTropa in niveles:
            nivel=niveles[tipoDeTropa]
            # Incremento de daño: +20% por nivel adicional
            tropa["Danio"]=round(base['Danio']*(1+0.2*(nivel-1)),2)
            # Incremento de vida: +25% por nivel adicional
            tropa["Vida"]=round(base['Vida']*(1+0.25*(nivel-1)),2)
            tropa["Nivel"]=nivel

#Funcion de mostrar los creditos
def creditos():

    global botonRegresarMenu
    
    #Intefaz de los creditos
    fondoMenu=pygame.image.load(imagenMenu)
    menuFondo=pygame.transform.scale(fondoMenu,(1280,720))
    fondoMenu=menuFondo

    fondoCreditosA=pygame.Rect(410,200,470,310)
    fondoCreditos=pygame.Rect(415,205,460,300)
    fondoDedicado=pygame.Rect(180,550,920,150)
    fondoDedicadoA=pygame.Rect(175,545,930,160)

    botonRegresarMenu=pygame.Rect(1070,10,200,70)
    botonRegresarMenuA=pygame.Rect(1065,5,210,80)
    fondoC=pygame.Rect(440,40,400,90)
    fondoCA=pygame.Rect(435,35,410,100)
    regresar=fuente1c.render("Regresar",True,colorTexto)

    titulo=fuente1a.render("Créditos",True,letras)
    hechoPor=fuente1b.render("Hecho por:",True,letras)
    nombres=[
        fuente1b.render("• Gabo",True,colorTexto),
        fuente1b.render("• Ahmed",True,colorTexto),
        fuente1b.render("• Maximo",True,colorTexto)
    ]

    dedicadoA=fuente1b.render("Dedicado al mejor profe de EDA:",True,letras)
    yovanni=fuente1b.render("Yovanni Álvarez Ulloa",True,colorTexto)

    #Objetos mostrados en la pantalla de 
    ventana.blit(fondoMenu,(0,0))
    pygame.draw.rect(ventana,cafeBorde,fondoCA,border_radius=10)
    pygame.draw.rect(ventana,cafeFondo,fondoC,border_radius=10)
    pygame.draw.rect(ventana,cafeBorde,fondoDedicadoA,border_radius=10)
    pygame.draw.rect(ventana,cafeFondo,fondoDedicado,border_radius=10)
    pygame.draw.rect(ventana,cafeBorde,fondoCreditosA,border_radius=30)
    pygame.draw.rect(ventana,cafeFondo,fondoCreditos,border_radius=30)
    ventana.blit(titulo,(460,43))
    ventana.blit(hechoPor,(500,230))
    for i, nombre in enumerate(nombres):
        ventana.blit(nombre,(445,300+i*60))
    ventana.blit(dedicadoA,(213,560))
    pygame.draw.rect(ventana,colorBorde,botonRegresarMenuA,border_radius=10)
    pygame.draw.rect(ventana,colorBoton,botonRegresarMenu,border_radius=10)
    ventana.blit(regresar,(1097,32))
    ventana.blit(yovanni,(350,620))

#Funcion del menú de partidas
def seleccionarpartida():
    
    #Variables que se convierten globales
    global partida1,partida2,partida3,partidab1,partidab2,partidab3
    global colorP1,colorP2,colorP3
    global nombre1,nombre2,nombre3
    global regresar2,borrar1,borrar2,borrar3

    #Carga fondo
    fondoMenu=pygame.image.load(imagenMenu)
    menuFondo=pygame.transform.scale(fondoMenu,(1280,720))
    fondoMenu=menuFondo
   
    #Obtiene el nombre del clan de un archivo o crea un archivo en caso de no existir 
    clan1=open(rutaClan1,"a+")
    clan1.seek(0)
    nombre1=clan1.readline().strip()
    clan1.close()
    if nombre1=="":
        nombre1="Crear partida"

    clan2=open(rutaClan2,"a+")
    clan2.seek(0)
    nombre2=clan2.readline().strip()
    clan2.close()
    if nombre2=="":
        nombre2="Crear partida"

    clan3=open(rutaClan3,"a+")
    clan3.seek(0)
    nombre3=clan3.readline().strip()
    clan3.close()
    if nombre3=="":
        nombre3="Crear partida"

    #Prepara los objetos que mostrara
    textoCargar=fuente1a.render("Cargar Partida",True,letras)
    P1=fuente1b.render("Partida 1",True,colorTexto)
    P2=fuente1b.render("Partida 2",True,colorTexto)
    P3=fuente1b.render("Partida 3",True,colorTexto)
    
    C1=fuente1b.render("Clan 1:",True,colorTexto)
    C2=fuente1b.render("Clan 2:",True,colorTexto)
    C3=fuente1b.render("Clan 3:",True,colorTexto)

    texto1=fuente1d.render(nombre1,True,blanco)
    texto2=fuente1d.render(nombre2,True,blanco)
    texto3=fuente1d.render(nombre3,True,blanco)
    textoBorrar=fuente1c.render("Borrar partida",True,negro)
    textoRegresar2=fuente1c.render("Regresar",True,colorTexto)
    b1=fuente1c.render("1",True,negro)
    b2=fuente1c.render("2",True,negro)
    b3=fuente1c.render("3",True,negro)

    partida1=pygame.Rect((95,150,300,420))
    partida2=pygame.Rect((490,150,300,420))
    partida3=pygame.Rect((885,150,300,420))
    partidab1=pygame.Rect((90,145,310,430))
    partidab2=pygame.Rect((485,145,310,430))
    partidab3=pygame.Rect((880,145,310,430))
    borrar=pygame.Rect(90,620,280,70)
    borrarA=pygame.Rect(85,615,290,80)
    borrar1=pygame.Rect(390,620,40,70)
    borrar1A=pygame.Rect(385,615,50,80)
    borrar2=pygame.Rect(440,620,40,70)
    borrar2A=pygame.Rect(435,615,50,80)
    borrar3=pygame.Rect(490,620,40,70)
    borrar3A=pygame.Rect(485,615,50,80)
    regresar2=pygame.Rect(990,620,200,70)
    regresar2A=pygame.Rect(985,615,210,80)
    FC=pygame.Rect(300,40,680,90)
    FCA=pygame.Rect(295,35,690,100)

    #Muestra los objetos 
    ventana.blit(fondoMenu,(0,0))
    pygame.draw.rect(ventana,colorBorde,partidab1,border_radius=10)
    pygame.draw.rect(ventana,colorBorde,partidab2, border_radius=10)
    pygame.draw.rect(ventana,colorBorde,partidab3, border_radius=10)
    pygame.draw.rect(ventana,colorBoton,partida1, border_radius=10)
    pygame.draw.rect(ventana,colorBoton,partida2, border_radius=10)
    pygame.draw.rect(ventana,colorBoton,partida3, border_radius=10)
    pygame.draw.rect(ventana,piedra,borrarA,border_radius=10)
    pygame.draw.rect(ventana,rojo,borrar,border_radius=10)
    pygame.draw.rect(ventana,colorBorde,regresar2A,border_radius=10)
    pygame.draw.rect(ventana,colorBoton,regresar2,border_radius=10)
    pygame.draw.rect(ventana,piedra,borrar1A,border_radius=10)
    pygame.draw.rect(ventana,piedra,borrar2A,border_radius=10)
    pygame.draw.rect(ventana,piedra,borrar3A,border_radius=10) 
    pygame.draw.rect(ventana,rojo,borrar1,border_radius=10)
    pygame.draw.rect(ventana,rojo,borrar2,border_radius=10)
    pygame.draw.rect(ventana,rojo,borrar3,border_radius=10) 
    pygame.draw.rect(ventana,cafeBorde,FCA,border_radius=10) 
    pygame.draw.rect(ventana,cafeFondo,FC,border_radius=10) 
 
    ventana.blit(textoCargar,(322,52))
    ventana.blit(textoBorrar,(102,640))
    ventana.blit(textoRegresar2,(1015,643))
    ventana.blit(b1,(402,640))
    ventana.blit(b2,(452,640))
    ventana.blit(b3,(502,640))
    
    ventana.blit(P1,(125,170))
    ventana.blit(P2,(520,170))
    ventana.blit(P3,(915,170))

    ventana.blit(C1,(125,300))
    ventana.blit(C2,(520,300))
    ventana.blit(C3,(915,300))

    ventana.blit(texto1,(130,450))
    ventana.blit(texto2,(535,450))
    ventana.blit(texto3,(930,450))

#Función para crear partida
def ingresarClan():
    
    global textoIngresado
    global regresar3

    #Prepara los objetos que mostrara
    fondoMenu=pygame.image.load(imagenMenu)
    menuFondo=pygame.transform.scale(fondoMenu,(1280,720))
    fondoMenu=menuFondo
    textoMostrar=fuente1b.render(textoIngresado, True, negro)
    texto=fuente1a.render("Ingresa el nombre de tu clan:",True,letras)
    ingresarR=pygame.Rect(30,85,1220,100)
    ingresarRA=pygame.Rect(25,80,1230,110)
    rect1=pygame.Rect(100,250,1080,160)
    rect2=pygame.Rect(110,260,1060,140)
    regresar3=pygame.Rect(990,620,200,70)
    regresar3A=pygame.Rect(985,615,210,80)
    textoRegresar2=fuente1c.render("Regresar",True,colorTexto)
    
    #Muestra los objetos
    ventana.blit(fondoMenu,(0,0))
    pygame.draw.rect(ventana,piedra,rect1,border_radius=10)
    pygame.draw.rect(ventana,gris,rect2,border_radius=10)
    pygame.draw.rect(ventana,colorBorde,regresar3A,border_radius=10)
    pygame.draw.rect(ventana,colorBoton,regresar3,border_radius=10)
    pygame.draw.rect(ventana,cafeBorde,ingresarRA,border_radius=10)
    pygame.draw.rect(ventana,cafeFondo,ingresarR,border_radius=10)
    ventana.blit(texto,(60,100))
    ventana.blit(textoMostrar, (150, 310))
    ventana.blit(textoRegresar2,(1015,643))

#Funcion que muestra la partida seleccionada
def iniciarJuego(clanIngresado):
    
    global botonJugar,regresar, botonPartidas
    
    #Crea una variable que contiene el documento de la partida seleccionada
    leer=os.path.join(Base,"clan"+str(clanIngresado)+".txt")

    #Abre el archivo para leerlo
    arch=open(leer,"r")
    clan=arch.readline().strip()
    arch.close()
    
    #Prepara que objetos mostrara
    fondoInicio=pygame.image.load(imagenInicio)
    fondoBien=pygame.transform.scale(fondoInicio,(1280,720))
    fondoInicio=fondoBien

    nombreClan=pygame.Rect(0,0,500,75)
    botonPartidas=pygame.Rect(0,80,500,75)
    bordeA=pygame.Rect(0,0,505,160)
    botonJugar=pygame.Rect(0,520,400,200)
    botonJugarA=pygame.Rect(0,515,405,210)
    regresar=pygame.Rect(1070,10,200,70)
    regresarA=pygame.Rect(1065,5,210,80)
    
    mostrarclan=fuente1c.render("Clan "+clan,True,letras)
    hist=fuente1c.render("Historial de partidas",True,letras)
    jugar=fuente1a.render("Jugar",True,letras)
    textoRegresar=fuente1c.render("Regresar",True,colorTexto)

    #Muestra los objetos
    ventana.blit(fondoInicio,(0,0))
    pygame.draw.rect(ventana,cafeBorde,bordeA,border_bottom_right_radius=10)
    pygame.draw.rect(ventana,cafeFondo,nombreClan,border_top_right_radius=10,border_bottom_right_radius=10)
    pygame.draw.rect(ventana,colorBoton,botonPartidas,border_top_right_radius=10,border_bottom_right_radius=10)
    pygame.draw.rect(ventana,colorBorde,botonJugarA,border_top_right_radius=50)
    pygame.draw.rect(ventana,colorBoton,botonJugar,border_top_right_radius=50)
    pygame.draw.rect(ventana,colorBorde,regresarA,border_radius=10)
    pygame.draw.rect(ventana,colorBoton,regresar,border_radius=10)
    
    ventana.blit(mostrarclan,(20,22))
    ventana.blit(hist,(75,102))
    ventana.blit(jugar,(70,585))        
    ventana.blit(textoRegresar,(1095,32)) 

#Funcion para elegir 3 guerreros        
def mostrarGuerreros():

    global botonVida,botonDanio,botonContinuar
    global botonRegresar
       
    #Declarar botones, rectangulos y texto
    fondoMenu=pygame.image.load(imagenGuerreros)
    
    botonVida=pygame.Rect(40,100,250,50)
    botonDanio=pygame.Rect(40,160,250,50)
    botonVidaA=pygame.Rect(35,95,260,60)
    botonDanioA=pygame.Rect(35,155,260,60)
    botonContinuar=pygame.Rect(1000,650,280,70)
    botonContinuarA=pygame.Rect(995,645,290,80)
    botonRegresar=pygame.Rect(1070,10,200,70)
    botonRegresarA=pygame.Rect(1065,5,210,80)
    fondoE=pygame.Rect(365,115,750,100)
    fondoEA=pygame.Rect(360,110,760,110)
    fondoO=pygame.Rect(5,23,370,60)
    fondoOA=pygame.Rect(0,18,380,70)

    textoVida=fuente1c.render("Vida",True,blanco)
    textoDanio=fuente1c.render("Daño",True,blanco)
    textoOrdenar=fuente1b.render("Ordenar por:",True,letras)    
    textoElegir=fuente1a.render("Elige 3 guerreros:",True,letras)
    textoContinuar=fuente1c.render("Continuar",True,blanco)
    textoRegresar=fuente1c.render("Regresar",True,colorTexto)

    #Mostrar los textos y rectangulos
    ventana.blit(fondoMenu,(0,0))
    pygame.draw.rect(ventana,piedra,botonVidaA,border_radius=10)
    pygame.draw.rect(ventana,piedra,botonDanioA,border_radius=10)
    pygame.draw.rect(ventana,cafeBorde,fondoEA,border_radius=10)
    pygame.draw.rect(ventana,cafeFondo,fondoE,border_radius=10)
    pygame.draw.rect(ventana,cafeBorde,fondoOA,border_radius=10)
    pygame.draw.rect(ventana,cafeFondo,fondoO,border_radius=10)
    pygame.draw.rect(ventana,verde,botonVida,border_radius=10)
    pygame.draw.rect(ventana,rojo,botonDanio,border_radius=10)
    pygame.draw.rect(ventana,colorBorde,botonRegresarA,border_radius=10)
    pygame.draw.rect(ventana,colorBoton,botonRegresar,border_radius=10)
    ventana.blit(textoVida,(130,110))
    ventana.blit(textoDanio,(130,170))
    ventana.blit(textoOrdenar,(5,30))
    ventana.blit(textoElegir,(380,130))
    ventana.blit(textoRegresar,(1095,32))

    #Condicion de continuar si ya eligio 3 guerreros
    if len(guerrerosElegidos)==3:
        pygame.draw.rect(ventana,piedra,botonContinuarA,border_top_left_radius=25)
        pygame.draw.rect(ventana,verde,botonContinuar,border_top_left_radius=25)
        ventana.blit(textoContinuar,(1060,675))
    
#Funcion dibujar tarjetas
def dibujarGuerreros(ventana,guerrero,x,y):

    #Condicion del color de las tarjetas si ya fueron seleccionadas
    if guerrero in guerrerosElegidos:
        colorFondo=piedra
    else:
        colorFondo=negro
   
    #Declaracion de los botones, textos y rectangulos
    nombre=fuente3a.render(guerrero["Nombre"],True,blanco)
    vida=fuente1c.render(f"Vida: {guerrero['Vida']}",True,blanco)
    danio=fuente1c.render(f"Daño: {guerrero['Danio']}",True,blanco)
    nivel=fuente1c.render(f"Nivel: {guerrero['Nivel']}",True,blanco)
    
    #Mostrar los objetos en pantalla
    borde=pygame.Rect(x,y,anchoTarjeta,altoTarjeta)
    pygame.draw.rect(ventana,colorBorde,borde,4,border_radius=10)
    pygame.draw.rect(ventana,colorFondo,(x+2,y+2,anchoTarjeta-4,altoTarjeta-4),border_radius=8)

    nombre_rect=nombre.get_rect(center=(x+anchoTarjeta/2,y+30))
    ventana.blit(nombre,nombre_rect)

    imagen=pygame.image.load(guerrero["imagenes"])
    tarjeta=imagen.get_rect(center=(x+anchoTarjeta/2,y+170))
    ventana.blit(imagen,tarjeta)

    ventana.blit(vida,(x+20,y+240))
    ventana.blit(danio,(x+20,y+265))
    ventana.blit(nivel,(x+20,y+300))
    
    return borde

#Funcion mostrar tarjetas y guerreros
def jugar():

    global cuadroGuerrero

    #Se ejecuta la funcion de mostrar los guerreros
    cuadroGuerrero=[]    
    mostrarGuerreros()
    
    #Se muestran las tarjetas en pantalla acomodandolas por posicion
    espacio=75
    tarjetas=anchoMenu//(anchoTarjeta+espacio)

    for i,guerrero in enumerate(guerreros):
        fila=i//tarjetas
        columna=i%tarjetas
        
        #Ordenar las tarjetas en la pantalla
        inicioX=10+espacio+columna*(anchoTarjeta+espacio)
        inicioY=200+espacio+fila*(altoTarjeta+espacio)
        
        cuadro=dibujarGuerreros(ventana,guerrero,inicioX,inicioY)
        cuadroGuerrero.append((cuadro,guerrero))

#Funcion de la cola de turnos
def colaTurnos1():
    global tamanioCola, colaActualizada, turnosVisibles, imagenesTurnos

    #Si la cola ya se actualizo, no se ejecuta la funcion
    if colaActualizada==True:
        return

    turnosVisibles=[]
    imagenesTurnos=[]
    
    #Se crea la cola de turnos con el tamaño de los guerreros
    tamanioCola=len(guerrerosElegidos)+len(enemigosRonda)

    #Se llena la cola con los guerreros elegidos y su color
    for i in range(tamanioCola):
        if i<len(guerrerosElegidos):
            cola.add_rear(("Azul", guerrerosElegidos[i]))
        if i<len(enemigosRonda):
            cola.add_rear(("Rojo", enemigosRonda[i]))

    turnos=min(tamanioCola, 6)

    #Se muestran los guerreros con su imagen cargada
    for i in range(turnos):
        jugador,guerrero=cola.remove_front()
        turnosVisibles.append((jugador, guerrero))

        imagen=pygame.image.load(guerrero["imagenes"])
        imagen=pygame.transform.scale(imagen,(200, 95))
        imagenesTurnos.append(imagen)

        cola.add_rear((jugador,guerrero))

    #Se actualiza la bandera de cola actualizada
    colaActualizada=True

#Funcion actualizar turnos y sus imagenes
def actualizarColaTurnos():
    global turnosVisibles,imagenesTurnos,c

    turnosVisibles=[]
    imagenesTurnos=[]

    #Se actualizan lsd imsgenes y se recorre la cola
    turnos=min(cola.size(),6)
    for i in range(turnos):
        jugador,guerrero=cola.remove_front()
        turnosVisibles.append((jugador,guerrero))

        imagen=pygame.image.load(guerrero["imagenes"])
        imagen=pygame.transform.scale(imagen,(200, 95))
        imagenesTurnos.append(imagen)

        cola.add_rear((jugador,guerrero))

#Funcion recursiva que llena las listas de vida, y elimina a los guerereros que ya murieron
def filtrarVivosRecursivo(listaPersonajes,listaVidas,i=0,nuevosPersonajes=None,nuevasVidas=None):
    
    #Primero evalua si ya hay listas creadas de los personajes y las vidas
    if nuevosPersonajes is None: 
        nuevosPersonajes=[]
    if nuevasVidas is None: 
        nuevasVidas=[]

    #Caso base, si la lista ya se lleno hasta el numero maximo, se rompe la recursividad y regresa las listas con los personajes y las vidas
    if i>=len(listaPersonajes):
        return nuevosPersonajes, nuevasVidas
    
    #Aqui evalua si todos los personajes estan vivos, si no lo estan, no los agrega a las listas
    if listaVidas[i]>0:
        nuevosPersonajes.append(listaPersonajes[i])
        nuevasVidas.append(listaVidas[i])

    #Regresa la funcion (Recursividad) para valuar cada indice de la lista hasta que las listas se llenen
    return filtrarVivosRecursivo(listaPersonajes,listaVidas,i+1,nuevosPersonajes,nuevasVidas)

#Funcion recursiva que regresa los enemigos que fueron utilizados en el combate
def reconstruirOrdenEnemigosRecursivo(enemigos,i=0,orden=None):
    
    #Se inicilaliza la lista al llamar la funcion
    if orden is None: 
        orden=[]

    #Caso base, si el indice es mayor que la cantidad de elementos que tiene la lista, se rompe la recursividad y regresa la lista
    if i>=len(enemigos):
        return orden
    
    #Si no se rompio la recursividad, se agrega a la lista el enemigo
    orden.append(enemigos[i]["Nombre"])

    #Regresa la funcion con un indice mas y la lista actualizada para irla recorriendo recursivamente
    return reconstruirOrdenEnemigosRecursivo(enemigos,i+1,orden)

#Funcion que evalua si alguna vida de algun guerrero es "0"
def ajustarVidasRecursivo(vidas,i=0,resultado=None):
    
    #Se inicializa la lista
    if resultado is None: 
        resultado=[]

    #Caso base, si el indice de la lista es mayor
    if i>=len(vidas):
        return resultado
    
    #Si no se rompio la recursividad, se agrega el valor de la vida pero con 0 si el valor de la vida es negativo (murio el guerrero)
    resultado.append(max(0,vidas[i]))

    #Se regresa la funcion recursiva con un indice mas y la lista del resultado
    return ajustarVidasRecursivo(vidas,i+1,resultado)

#Funcion para actualizar el estado del combate despues de un turno, verificando todos los casos de recursividad
def actualizarEstadoCombate():
    global guerrerosElegidos,vidasGuerreros
    global enemigosRonda,vidasEnemigos,ordenEnemigos
    global cola,colaActualizada

    #Se llaman a las funciones recursivas para reordenar las listas de las vidas, guerreros y guerreros muertos
    guerrerosElegidos,vidasGuerreros=filtrarVivosRecursivo(guerrerosElegidos,vidasGuerreros)
    enemigosRonda,vidasEnemigos=filtrarVivosRecursivo(enemigosRonda,vidasEnemigos)
    ordenEnemigos[:]=reconstruirOrdenEnemigosRecursivo(enemigosRonda)
    vidasGuerreros=ajustarVidasRecursivo(vidasGuerreros)
    vidasEnemigos=ajustarVidasRecursivo(vidasEnemigos)

    #Se guarda el tamaño de la cola para actualizarlo con el ciclo for
    tam=cola.size()
    for i in range(tam):
        jugador,guerrero=cola.remove_front()
        vida=0

        if jugador=="Azul":
            # Buscar índice del guerrero en la lista de los guerreros elegidos
            indice1=-1
            for i in range(len(guerrerosElegidos)):
                if guerrerosElegidos[i] == guerrero:
                    indice1=i
                    break
            if indice1!=-1:
                vida=vidasGuerreros[indice1]

        else:  
            indice1=-1
            for i in range(len(enemigosRonda)):
                if enemigosRonda[i]==guerrero:
                    indice1=i
                    break
            if indice1!=-1:
                vida=vidasEnemigos[indice1]

        if vida>0:
            cola.add_rear((jugador,guerrero))
        
        #Evalua cada vida de la cola de turnos, y si la vida es menor o igual a 0, o sea esta muerto, no se vuelve a ingresar a la cola

#Funcion de combate completo
def combate():
    
    global ataqueDanio,curarVida
    global guerrerosEnemigos
    global vidasEnemigos,vidasGuerreros
    global enemigosRonda
    global atacar1,atacar2,atacar3,textoAtacar1,textoAtacar2,textoAtacar3

    #Se declara todo lo que se va a mostrar
    fondoBatalla=pygame.image.load(campoBatalla)
    campoFondo=pygame.transform.scale(fondoBatalla,(1280,720))
    fondoBatalla=campoFondo
    
    turnos=pygame.Rect(0,0,1280,100)
    turnosA=pygame.Rect(0,0,1280,105)
    ataques=pygame.Rect(0,570,1280,150)
    ataquesA=pygame.Rect(0,565,1280,155)
    turno1=pygame.Rect(350,12,150,75)
    turno2=pygame.Rect(600,12,150,75)
    turno3=pygame.Rect(850,12,150,75)
    turno4=pygame.Rect(1100,12,150,75)
    turno1A=pygame.Rect(345,7,160,85)
    turno2A=pygame.Rect(595,7,160,85)
    turno3A=pygame.Rect(845,7,160,85)
    turno4A=pygame.Rect(1095,7,160,85)
    ataqueDanio=pygame.Rect(200,600,400,100)
    ataqueDanioA=pygame.Rect(195,595,410,110)
    curarVida=pygame.Rect(660,600,400,100)
    curarVidaA=pygame.Rect(655,595,410,110)
    colaTurnos=fuente1c.render("Cola de turnos:",True,colorTexto)
    atacar=fuente1b.render("Atacar",True,negro)
    curar=fuente1b.render("Curar +20",True,negro)

    textoAtacar1=fuente1c.render("Elegir",True,negro)
    textoAtacar2=fuente1c.render("Elegir",True,negro)
    textoAtacar3=fuente1c.render("Elegir",True,negro)

    #Se muestran las vidas de los guerreros, y si uno muere, se muestra una "X"
    vida1=str(round(vidasGuerreros[0],2)) if len(vidasGuerreros)>0 else "X"
    vida2=str(round(vidasGuerreros[1],2)) if len(vidasGuerreros)>1 else "X"
    vida3=str(round(vidasGuerreros[2],2)) if len(vidasGuerreros)>2 else "X"
    vida4= str(round(vidasEnemigos[0],2)) if len(vidasEnemigos)>0 else "X"
    vida5=str(round(vidasEnemigos[1],2)) if len(vidasEnemigos)>1 else "X"
    vida6=str(round(vidasEnemigos[2],2)) if len(vidasEnemigos)>2 else "X"
    
    textoVida1=fuente1c.render(vida1,True,verde)
    textoVida2=fuente1c.render(vida2,True,verde)
    textoVida3=fuente1c.render(vida3,True,verde)
    textoVida4=fuente1c.render(vida4,True,verde)
    textoVida5=fuente1c.render(vida5,True,verde)
    textoVida6=fuente1c.render(vida6,True,verde)

    cuadroV1=pygame.Rect(270,140,90,50)
    cuadroV2=pygame.Rect(430,240,90,50)
    cuadroV3=pygame.Rect(270,490,90,50)
    cuadroV4=pygame.Rect(890,140,90,50)
    cuadroV5=pygame.Rect(710,260,90,50)
    cuadroV6=pygame.Rect(890,490,90,50)
    
    atacar1=pygame.Rect(730,150,150,50)
    atacar2=pygame.Rect(580,330,150,50)
    atacar3=pygame.Rect(730,480,150,50)
    atacar1A=pygame.Rect(725,145,160,60)
    atacar2A=pygame.Rect(575,325,160,60)
    atacar3A=pygame.Rect(725,475,160,60)

    cuadroTurnos=pygame.Rect(400,100,480,30)
    cuadroTurnosA=pygame.Rect(395,95,490,40)

    cuadroT=pygame.Rect(400,545,480,30)
    cuadroTA=pygame.Rect(395,540,490,40)

    guerrerosCargados=[]
    guerrerosEnemigos=[]
    guerrerosCargadosCola=[]
    guerrerosEnemigosCola=[]

    #Agrega a la lista guerreros cargados, las imagenes de los guerreros seleccionados
    for i,guerreros in enumerate(guerrerosElegidos):
        imagenCargada=pygame.image.load(guerreros["imagenes"])
        guerrerosCargados.append(imagenCargada)
   
    #Agrega a la lista guerrerosEnemigos, las imagenes de los enemigos que apareceran
    for i,enemigos in enumerate(enemigosRonda):
        enemigoCargado=pygame.image.load(enemigos["imagenes"])
        guerrerosEnemigos.append(enemigoCargado)
    
    for i, guerreros in enumerate(guerrerosElegidos):
        imagenCargada=pygame.image.load(guerreros["imagenes"])
        imagenRedimensionada=pygame.transform.scale(imagenCargada,(200,95))
        guerrerosCargadosCola.append(imagenRedimensionada)
    
    for i, guerreros in enumerate(enemigosRonda):
        imagenCargada=pygame.image.load(enemigos["imagenes"])
        imagenRedimensionada=pygame.transform.scale(imagenCargada,(200,95))
        guerrerosEnemigosCola.append(imagenRedimensionada)

    
    #Se muestran en pantalla los botones, rectangulos y textos
    ventana.blit(fondoBatalla,(0,0))
    pygame.draw.rect(ventana,colorBorde,cuadroTurnosA,border_radius=10)
    pygame.draw.rect(ventana,colorBorde,turnosA,border_bottom_left_radius=20,border_bottom_right_radius=20)
    pygame.draw.rect(ventana,colorBorde,ataquesA,border_top_left_radius=20,border_top_right_radius=20)
    pygame.draw.rect(ventana,cafeFondo,turnos,border_bottom_left_radius=20,border_bottom_right_radius=20)
    pygame.draw.rect(ventana,cafeBorde,cuadroTA,border_radius=10)
    pygame.draw.rect(ventana,cafeFondo,cuadroT,border_radius=10)
    pygame.draw.rect(ventana,cafeFondo,ataques,border_top_left_radius=20,border_top_right_radius=20)
    pygame.draw.rect(ventana,cafeFondo,cuadroTurnos,border_bottom_left_radius=10, border_bottom_right_radius=10)
    pygame.draw.rect(ventana,piedra,turno1A,border_radius=20)
    pygame.draw.rect(ventana,piedra,turno2A,border_radius=20)
    pygame.draw.rect(ventana,piedra,turno3A,border_radius=20)
    pygame.draw.rect(ventana,piedra,turno4A,border_radius=20)    
    pygame.draw.rect(ventana,blanco,turno1,border_radius=20)
    pygame.draw.rect(ventana,blanco,turno2,border_radius=20)
    pygame.draw.rect(ventana,blanco,turno3,border_radius=20)
    pygame.draw.rect(ventana,blanco,turno4,border_radius=20)

    #Se usa una variable para recuperar el siguiente turno que esta en la cola
    bandera=cola.remove_front()
    cola.add_front(bandera)

    #Se imprime un texto con el siguiente turno
    turnosTexto="Turno de: "+bandera[1]["Nombre"]+" "+bandera[0]
    textoImprimir=fuente1c.render(turnosTexto,True,colorTexto)
    
    #Botones y textos del turno del equipo azul
    if bandera[0]=="Azul":
        pygame.draw.rect(ventana,piedra,ataqueDanioA,border_radius=20)
        pygame.draw.rect(ventana,piedra,curarVidaA,border_radius=20)
        pygame.draw.rect(ventana,rojo,ataqueDanio,border_radius=20)
        pygame.draw.rect(ventana,verde,curarVida,border_radius=20)
        ventana.blit(atacar,(305, 630))
        ventana.blit(curar,(735,630))

    if elegirAtaque:
        if len(vidasEnemigos)>0:
            pygame.draw.rect(ventana,piedra,atacar1A,border_radius=10)
            pygame.draw.rect(ventana,amarillo,atacar1,border_radius=10)
            ventana.blit(textoAtacar1,(755,162))
            
        if len(vidasEnemigos)>1:
            pygame.draw.rect(ventana,piedra,atacar2A,border_radius=10)
            pygame.draw.rect(ventana,amarillo,atacar2,border_radius=10)
            ventana.blit(textoAtacar2,(605,342))
        if len(vidasEnemigos)>2:
            pygame.draw.rect(ventana,piedra,atacar3A,border_radius=10)
            pygame.draw.rect(ventana,amarillo,atacar3,border_radius=10)
            ventana.blit(textoAtacar3,(755,492))
        
    #Ciclos for para imprimir las imagenes de los personajes en la cola de turnos
    for i, imagen in enumerate(imagenesTurnos):
        x=320+250*i
        y=-1
        ventana.blit(imagen,(x,y))

    #Imprime al equipo del jugador 
    for i in range(len(guerrerosCargados)):
        if i%2!=0:
            x=-200+200*i
        else:
            x=-200
        y=50+100*i        
        ventana.blit(guerrerosCargados[i],(x,y))
    
    #Imprime a los enemigos
    for i in range(len(guerrerosEnemigos)):
        if i%2!=0:
            x=400+150*i
        else:
            x=750
        y=75+100*i
        ventana.blit(guerrerosEnemigos[i],(x,y))

    #Se imprimen los rectangulos, botones y texto    
    pygame.draw.rect(ventana,negro,cuadroV1,border_radius=15)
    pygame.draw.rect(ventana,negro,cuadroV2,border_radius=15)
    pygame.draw.rect(ventana,negro,cuadroV3,border_radius=15)
    pygame.draw.rect(ventana,negro,cuadroV4,border_radius=15)
    pygame.draw.rect(ventana,negro,cuadroV5,border_radius=15)
    pygame.draw.rect(ventana,negro,cuadroV6,border_radius=15)
    
    
    ventana.blit(textoVida1,(282,152))
    ventana.blit(textoVida2,(442,252))
    ventana.blit(textoVida3,(282,502))
    ventana.blit(textoVida4,(902,152))
    ventana.blit(textoVida5,(722,272))
    ventana.blit(textoVida6,(902,502))

    ventana.blit(textoImprimir,(420,100))
    
    ventana.blit(colaTurnos,(60,35))

#Funcion de avidos para elegir si es mejor curarse o atacar
def ataqueEvaluarAvidos1(colaTurnos):
    
    global yaAtaco

    #Se abre el archivo de admin para verificar cual es la dificultad
    admi=open(rutaAdmin,"r")
    p=admi.readline()
    dificultad=admi.readline()

    turno=colaTurnos.remove_front()
    colaTurnos.add_front(turno)

    #La funcion no se ejecuta si es turno del azul
    if turno[0]!="Rojo":
        return

    #Busca de que enemigo es turno
    indice=next((i for i, n in enumerate(ordenEnemigos) if n==turno[1]["Nombre"]),-1)
    if indice==-1:
        return

    #Obtiene la vida maxima de ese enemigo y la que le queda
    vidaActual=vidasEnemigos[indice]
    vidaMaxima=turno[1]["Vida"]

    acciones=[]

    #Segun la dificultad seleccionada en admin
    if dificultad=="Fácil":
        #En facil se curan los enemigos que tienen menos del 40% de vida
        if vidaActual<0.4*vidaMaxima:
            acciones.append(("curar",vidaMaxima-vidaActual))
    elif dificultad=="Difícil":
        #En dificil se curan los enemigos que tienen menos del 70% de vida
        if vidaActual<0.7*vidaMaxima:
            acciones.append(("curar",vidaMaxima-vidaActual))

    #Se ingresa en una lista la accion elegida y un valor para comparar

    acciones.append(("atacar", max(guerrero["Danio"] for i, guerrero in enumerate(guerrerosElegidos) if vidasGuerreros[i]>0)))

    #Avidos
    accion=sorted(acciones,key=lambda x: x[1],reverse=True)[0][0]
    #Se evalua cual es la mejor opcion, si atacar o curarse
    #Si se curan, se aumenta un 20 de vida    
    if accion=="curar":
        vidasEnemigos[indice]=min(vidaMaxima,vidaActual+20)
        mensaje="Acción: "+turno[1]["Nombre"]+" Rojo se curó"
    else:
        #Si no se curan, se usa avidos para evaluar el mejor ataque
        objetivo=seleccionarMejorObjetivoAvidos2()
        #Se elige el mejor ataque y se le resta el valor del daño del enemigo
        if objetivo is not None:
            vidasGuerreros[objetivo]-=turno[1]["Danio"]
            vidasGuerreros[objetivo]=max(0,vidasGuerreros[objetivo])
            mensaje="Acción: "+turno[1]["Nombre"]+" Rojo atacó a "+guerrerosElegidos[objetivo]["Nombre"]+" Azul"

    #Se actualiza la cola de turnos
    actualizarEstadoCombate()
    colaTurnos.add_rear(colaTurnos.remove_front())
    actualizarColaTurnos()

    #Se muestra en pantalla la accion realizada
    textoAccion=fuente1d.render(mensaje,True,colorTexto)
    pygame.display.update() 
    ventana.blit(textoAccion,(405,550))
    pygame.display.update()
    pygame.time.delay(3500)

    combate()
    pygame.display.update()
    pygame.time.delay(500)

    yaAtaco=False

#Funcion con avidos que valua cual es el mejor personaje para atacar segun su vida
def seleccionarMejorObjetivoAvidos2():
    #Se inicializa el indice del mejor objetivo a atacar y un valor de que tan buen ataque es
    mejorIndice=None
    mejorValor=-1
    for i, guerrero in enumerate(guerrerosElegidos):
        if vidasGuerreros[i]<=0:
            continue
        #Va recorriendo la lista de los guerreros y va calculando que tan buen ataque es
        cambioDeVida=vidasGuerreros[i]/guerrero["Vida"]
        valor=guerrero["Danio"]*(1-cambioDeVida)
        #Va actualizando el mejor valor y el mejor indice para atacar
        if valor>mejorValor:
            mejorValor=valor
            mejorIndice=i
    #Regresa el mejor indice de la lista de aliados
    return mejorIndice

#Funcion para curar a un guerreo        
def curar(colaTurnos):

    turno=colaTurnos.remove_front()
    #enemigosRonda=("")
    #turno=(("Color",diccionario del guerrero))
    colaTurnos.add_front(turno)
    #vidasEnemigos(70,150,100)
    #ordenEnemigos("Arquero","Gigante","Caballero")    
    
    #Si es el turno del azul, se curara si se aprieta el boton
    if turno[0]=="Azul":
        for i in range(len(vidasGuerreros)):
            if guerrerosElegidos[i]["Nombre"]==turno[1]["Nombre"] and turno[1]["Vida"]==vidasGuerreros[i]:
                textoCurar=fuente3a.render("Vida al máximo",True,colorTexto)
                pygame.display.update() 
                ventana.blit(textoCurar,(405,550))
                pygame.display.update()
                pygame.time.delay(1500)
                vidaMaxima=True
            elif guerrerosElegidos[i]["Nombre"]==turno[1]["Nombre"] and turno[1]["Vida"]!=vidasGuerreros[i]:
                vidasGuerreros[i]+=20
                if vidasGuerreros[i]>turno[1]["Vida"]:
                    vidasGuerreros[i]=turno[1]["Vida"]
                print("aliado curado")
                print("enemigo vida: ",vidasGuerreros[i])
        
                colaTurnos.add_rear(colaTurnos.remove_front())
                yaAtaco=False
     
    else:
        return

#Funcion para vrerificar que no se sobreescriba el archivo excel
def verificarExcel():
    
    #Si no existe el archivo de excel, lo crea y lo inicializa con los datos a guardar y las hojas
    try:
        historialExcel=openpyxl.load_workbook(rutaExcel)
        print("el archivo ya existe")
    except FileNotFoundError:
        historialExcel=openpyxl.Workbook()
        hojaClan1=historialExcel.active
        hojaClan1.title="HistorialClan1"
        hojaClan2=historialExcel.create_sheet("HistorialClan2")
        hojaClan3=historialExcel.create_sheet("HistorialClan3")
        
        encabezados=["Partida","Guerrero1","Guerrero2","Guerrero3","Resultado"]
        hojaClan1.append(encabezados)
        hojaClan2.append(encabezados)
        hojaClan3.append(encabezados)

        hojaClan1.cell(row=2,column=1).value=0
        hojaClan2.cell(row=2,column=1).value=0
        hojaClan3.cell(row=2,column=1).value=0

        historialExcel.save(rutaExcel)  
        historialExcel.close()

        print("archivo creado")

#Funcion historial de partidas con excel y listas
def historial():
    
    global colaHistorial
    
    historialExcel=openpyxl.load_workbook(rutaExcel)
    
    #Se carga la hoja de excel segun la partida cargada
    if clanIngresado==1:
        hojaClan=historialExcel["HistorialClan1"]
    elif clanIngresado==2:
        hojaClan=historialExcel["HistorialClan2"]
    elif clanIngresado==3:
        hojaClan=historialExcel["HistorialClan3"]

    #Se obtiene el tamaño de las columnas, que es igual a las partidas jugadas    
    tamanioHistorial=0   
    for i in hojaClan['A']:
        if i is not None:    
            tamanioHistorial+=1    

    colaHistorial=Deque()

    #Se llena la lista con los datos del excel    
    for i in range(2,tamanioHistorial+1):
        diccionarioRegistro={"Partida":hojaClan.cell(row=i,column=1).value,
                                "Guerrero1":hojaClan.cell(row=i,column=2).value,
                                "Guerrero2":hojaClan.cell(row=i,column=3).value,
                                "Guerrero3":hojaClan.cell(row=i,column=4).value,
                                "Resultado":hojaClan.cell(row=i,column=5).value}
        colaHistorial.add_rear(diccionarioRegistro)
    
    print(colaHistorial._items)

#Funcion borrar la partida y sus datos
def borrarPartidaExcel():

    #Se abre el archivo y se obtiene la hoja
    historialExcel=openpyxl.load_workbook(rutaExcel)

    if clanIngresado==1:
        hojaClan=historialExcel["HistorialClan1"]
    elif clanIngresado==2:
        hojaClan=historialExcel["HistorialClan2"]
    elif clanIngresado==3:
        hojaClan=historialExcel["HistorialClan3"]

    #Se obtiene el tamaño de las columnas, que es igual a las partidas jugadas    
    tamanioHistorial=0   
    for i in hojaClan['A']:
        if i is not None:    
            tamanioHistorial+=1    

    #Se llena la lista con los datos del excel    
    for i in range(2,tamanioHistorial+1):
        hojaClan.cell(row=i,column=1).value=None
        hojaClan.cell(row=i,column=2).value=None
        hojaClan.cell(row=i,column=3).value=None
        hojaClan.cell(row=i,column=4).value=None
        hojaClan.cell(row=i,column=5).value=None

    historialExcel.save(rutaExcel)
    historialExcel.close()
    print("partida borrada",clanIngresado)

#Funcion que muestra la interfza del historial de partidas
def mostrarInterfazHistorial(): 
    
    global botonRegresarH,arriba,abajo

    #Cargamos las imagenes, botones y textos
    fondoHistorial=pygame.image.load(imagenHistorial)
    historialImagen=pygame.transform.scale(fondoHistorial,(1280,720))
    fondoHistorial=historialImagen

    textoHistorial=fuente1b.render("Historial de partidas",True,letras)
    textoRegresar=fuente1c.render("Regresar",True,colorTexto)

    botonRegresarHA=pygame.Rect(500,595,280,90)
    botonRegresarH=pygame.Rect(505,600,270,80)
    mostrar=pygame.Rect(300,50,680,90)
    mostrarA=pygame.Rect(295,45,690,100)
    flechaArriba=pygame.image.load(imagenFlechaArriba)
    flechaAbajo=pygame.image.load(imagenFlechaAbajo)
    arribaFlecha=pygame.transform.scale(flechaArriba,(50,50))
    abajoFlecha=pygame.transform.scale(flechaAbajo,(50,50))

    #Mostramos en pantalla
    ventana.blit(fondoHistorial,(0,0))
    pygame.draw.rect(ventana,colorBorde,botonRegresarHA,border_radius=10)
    pygame.draw.rect(ventana,colorBoton,botonRegresarH,border_radius=10)
    pygame.draw.rect(ventana,cafeBorde,mostrarA,border_radius=10)
    pygame.draw.rect(ventana,cafeFondo,mostrar,border_radius=10)
    ventana.blit(textoHistorial,(340,75))
    ventana.blit(textoRegresar,(568,627))

    arriba=pygame.Rect(980,345,50,50)
    abajo=pygame.Rect(980,415,50,50)
    
    pygame.draw.rect(ventana,colorBoton,arriba,border_radius=10)
    pygame.draw.rect(ventana,colorBoton,abajo,border_radius=10)

    ventana.blit(arribaFlecha,(980,345))
    ventana.blit(abajoFlecha,(978,415))

    coordenadasPantalla=[230,330,425]
    
    #Se muestran las 3 partidas cargadas en la lista de partidas que fueron obtenidos de la cola de partidas
    for i,partida in enumerate(partidas):
        numero="Partida "+str(partida["Partida"])+":"
        guerreros=[partida["Guerrero1"],partida["Guerrero2"],partida["Guerrero3"]]
        resultado=partida["Resultado"]

        coordenada=coordenadasPantalla[i]
        ventana.blit(fuente1c.render(numero,True,negro),(330,coordenada+45))
        for posicion,guerrero in enumerate(guerreros):
            ventana.blit(fuente1d.render(guerrero,True,negro),(340+posicion*120,coordenada+85))
        if resultado=="Victoria":
            ventana.blit(fuente1b.render(resultado,True,verde),(715,coordenada+50))
        else:
            ventana.blit(fuente1b.render(resultado, True, rojo),(715,coordenada+50))

#Función para el historial de partidas(arriba)
def cargarColaHistorial():
    global partidas
    partidas=[]
    
    #Se recorre la cola y se actualiza la lista de las ultimas 3 partidas
    if colaHistorial.size()==0:

        return

    for i in range(min(3,colaHistorial.size())):
        ultimaPartida=colaHistorial.remove_rear()
        partidas.append(ultimaPartida)
      
    for ultimaPartida in partidas:
        colaHistorial.add_front(ultimaPartida)

#Función para el historial de partidas(abajo)
def cargarColaHistorialAlreves():
    global partidas
    partidas=[]

    #Se recorre la cola alreves y se actualiza la lista de las ultimas partidas
    if colaHistorial.size()==0:
        return

    for i in range(min(3,colaHistorial.size())):
        ultimaPartida=colaHistorial.remove_front()
        partidas.append(ultimaPartida) 
      
    for ultimaPartida in partidas:
        colaHistorial.add_rear(ultimaPartida)

    partidas=list(reversed(partidas))

#Funcion que muestra el resultado de al momento de terminar una partida
def mostrarResultado(resultado):

    global estado
    global enemigosDB, guerrerosRojos

    #Se imprime el resultado de la partida y se actualizan los niveles
    #Para cada caso de actualizan los archivos de los niveles, recuperando la lista de los guerreros y usando las funciones para mejorar guerreros
    if resultado=="Victoria":
        
        textoVictoria=fuente1.render("VICTORIA",True,verdeOscuro)
        ventana.blit(textoVictoria,(270,270))
        if clanIngresado==1:
            archivo=open(rutaClan1,"r")  
            lineas = archivo.readlines()
            archivo.close()
            nuevoTexto= []
            for linea in lineas:
                if ":" not in linea:
                    nuevoTexto.append(linea)
                    continue
                nombre, nivel = linea.strip().split(":")
                nombre = nombre.strip()
                nivel = int(nivel.strip())
                for tropa in copiaGuerreros:
                    if tropa["Nombre"] == nombre:
                        nivel += 1
                        break
                nuevoTexto.append(nombre+": "+str(nivel)+"\n")
            archivo=open(rutaClan1, "w")
            archivo.writelines(nuevoTexto)
            archivo.close()
            mejorarPersonajes(guerreros,guerrerosDB,rutaClan1)
            mejorarEnemigos(guerrerosRojos,enemigosDB,rutaClan1)
            
        elif clanIngresado==2:
            archivo=open(rutaClan2,"r")  
            lineas = archivo.readlines()
            archivo.close()
            nuevoTexto= []
            for linea in lineas:
                if ":" not in linea:
                    nuevoTexto.append(linea)
                    continue
                nombre, nivel = linea.strip().split(":")
                nombre = nombre.strip()
                nivel = int(nivel.strip())
                for tropa in copiaGuerreros:
                    if tropa["Nombre"] == nombre:
                        nivel += 1
                        break
                nuevoTexto.append(nombre+": "+str(nivel)+"\n")
            archivo=open(rutaClan2, "w")
            archivo.writelines(nuevoTexto)
            archivo.close()
            mejorarPersonajes(guerreros,guerrerosDB,rutaClan2)
            mejorarEnemigos(guerrerosRojos,enemigosDB,rutaClan2)
            
        elif clanIngresado==3:
            archivo=open(rutaClan3,"r")  
            lineas = archivo.readlines()
            archivo.close()
            nuevoTexto= []
            for linea in lineas:
                if ":" not in linea:
                    nuevoTexto.append(linea)
                    continue
                nombre, nivel = linea.strip().split(":")
                nombre = nombre.strip()
                nivel = int(nivel.strip())
                for tropa in copiaGuerreros:
                    if tropa["Nombre"] == nombre:
                        nivel += 1
                        break
                nuevoTexto.append(nombre+": "+str(nivel)+"\n")
            archivo=open(rutaClan3, "w")
            archivo.writelines(nuevoTexto)
            archivo.close() 
            mejorarPersonajes(guerreros,guerrerosDB,rutaClan3)
            mejorarEnemigos(guerrerosRojos,enemigosDB,rutaClan3)
            

    elif resultado=="Derrota":    
        textoDerrota=fuente1.render("DERROTA",True,rojo)
        ventana.blit(textoDerrota,(270,270))
 
    historialExcel=openpyxl.load_workbook(rutaExcel)
    
    #Se carga la hoja de excel segun la partida cargada
    if clanIngresado==1:
        hojaClan=historialExcel["HistorialClan1"]
    elif clanIngresado==2:
        hojaClan=historialExcel["HistorialClan2"]
    elif clanIngresado==3:
        hojaClan=historialExcel["HistorialClan3"]

    #Se obtiene el tamaño de las columnas, que es igual a las partidas jugadas    
    tamanioHistorial=0   
    for i in hojaClan['A']: 
        if i is not None:    
            tamanioHistorial+=1
        else:
            break

    #Se guarda en la hoja de excel el resultado de la partida
    hojaClan.cell(row=tamanioHistorial,column=5,value=resultado)  

    historialExcel.save(rutaExcel)
    historialExcel.close()

    copiaGuerreros.clear()
    enemigosRonda.clear()
    guerrerosElegidos.clear()
    vidasEnemigos.clear()
    vidasGuerreros.clear()
    turnosVisibles.clear()
    imagenesTurnos.clear()
    ordenEnemigos.clear()
    while cola.size()!=0:
        cola.remove_front()    
    

    pygame.display.update()
    pygame.time.delay(2000)

#Funciones de quicksort    
def guerrerosQuicksort(listaGuerreros): 
    quicksort_aux(listaGuerreros, 0, len(listaGuerreros) - 1)

def quicksort_aux(lista, inicio, fin):
    if inicio < fin:
        pivote = particion(lista, inicio, fin)
        quicksort_aux(lista, inicio, pivote - 1)
        quicksort_aux(lista, pivote + 1, fin)

def particion(lista, inicio, fin):
    
    pivote = lista[inicio]
    izquierda = inicio + 1
    derecha = fin

    bandera = False
    while not bandera:

        while (izquierda <= derecha and
                (lista[izquierda][sortear]>pivote[sortear] or (lista[izquierda][sortear]==pivote[sortear] and lista[izquierda][sortear]>=pivote[sortear]))):
                izquierda = izquierda + 1
                
        while (derecha >= izquierda and (lista[derecha][sortear] < pivote[sortear] or (lista[derecha][sortear] == pivote[sortear] and lista[derecha][sortear] <= pivote[sortear]))):
            derecha = derecha - 1
            
        if derecha < izquierda:
            bandera = True
        else:
            temp = lista[izquierda]
            lista[izquierda] = lista[derecha]
            lista[derecha] = temp

    temp = lista[inicio]
    lista[inicio] = lista[derecha]
    lista[derecha] = temp
    
    return derecha

#Ciclo while donde se programan los botones y teclado, la interacción con el usuario
ejecutando=True
while ejecutando:
    #Caso base si se da click en el tache de la ventana para cerrar la aplicacion
    for event in pygame.event.get():
        if event.type == pygame.QUIT:   
            ejecutando = False

        #Botones de la pantalla de menu
        if estado=="menu":
            if event.type == pygame.MOUSEBUTTONDOWN:
                if boton1.collidepoint(event.pos):
                    estado="seleccionarPartida"
                elif botonCreditos.collidepoint(event.pos):
                    estado="creditos"
                elif botonSalir.collidepoint(event.pos):
                    ejecutando=False
                elif botonAdmin.collidepoint(event.pos):
                    estado="contraseniaAdmin"
        
        #Botones e inicializacion y carga de los archivos al cargar una partida
        elif estado=="seleccionarPartida":
            if event.type == pygame.MOUSEBUTTONDOWN:
                if partida1.collidepoint(event.pos):
                    verificarExcel()
                    colorP1=piedra
                    clanIngresado=1
                    if nombre1=="Crear partida":
                        estado="ingresarClan"
                    else:
                        mejorarPersonajes(guerreros,guerrerosDB,rutaClan1)
                        mejorarEnemigos(guerrerosRojos,enemigosDB,rutaClan1)
                        estado="iniciarJuego"
                elif partida2.collidepoint(event.pos):
                    verificarExcel()
                    colorP2=piedra
                    clanIngresado=2
                    if nombre2=="Crear partida":
                        estado="ingresarClan"
                    else:
                        mejorarPersonajes(guerreros,guerrerosDB,rutaClan2)
                        mejorarEnemigos(guerrerosRojos,enemigosDB,rutaClan2)
                        estado="iniciarJuego"
                elif partida3.collidepoint(event.pos):
                    verificarExcel()
                    colorP3=piedra
                    clanIngresado=3
                    if nombre3=="Crear partida":
                        estado="ingresarClan"
                    else:
                        mejorarPersonajes(guerreros,guerrerosDB,rutaClan3)
                        mejorarEnemigos(guerrerosRojos,enemigosDB,rutaClan3)
                        estado="iniciarJuego"
                elif regresar2.collidepoint(event.pos):
                    estado="menu"
                elif borrar1.collidepoint(event.pos):
                    clanIngresado=1
                    clan1=open(rutaClan1,"w")
                    clan1.write("")
                    clan1.close()
                    textoIngresado=""
                    borrarPartidaExcel()
                elif borrar2.collidepoint(event.pos):
                    clanIngresado=2
                    clan2=open(rutaClan2,"w")
                    clan2.write("")
                    clan2.close()
                    textoIngresado="" 
                    borrarPartidaExcel()
                elif borrar3.collidepoint(event.pos):
                    clanIngresado=3
                    clan3=open(rutaClan3,"w")
                    clan3.write("")
                    clan3.close()
                    textoIngresado=""
                    borrarPartidaExcel()

        #Teclado para guardar el texto ingresado como nombre del clan en los archivos de cada clan        
        elif estado=="ingresarClan":
            if event.type == pygame.MOUSEBUTTONDOWN:
                if regresar3.collidepoint(event.pos):
                    estado="seleccionarPartida"
            if event.type == pygame.KEYDOWN:
                if event.key == pygame.K_RETURN:
                    print("Clan a guardar: ",clanIngresado," ,Nombre: ",textoIngresado)
                    if len(textoIngresado) <= 0 or len(textoIngresado) > 15:
                        mensajeError = fuente2.render("El nombre debe tener entre 1 y 15 caracteres", True, rojo)
                        ventana.blit(mensajeError, (100, 400))
                        pygame.display.update()
                        pygame.time.delay(2000)
                    else:
                        textoIngresado=textoIngresado+"\n"
                        if clanIngresado==1:
                            clan1=open(rutaClan1,"w+")
                            clan1.write(textoIngresado)
                            clan1.write("Espadachin: 1\n")
                            clan1.write("Arquero: 1\n")
                            clan1.write("Gigante: 1\n")
                            clan1.write("Caballero: 1\n")
                            clan1.close()
                                 
                        elif clanIngresado==2:
                            clan2=open(rutaClan2,"w+")
                            clan2.write(textoIngresado)
                            clan2.write("Espadachin: 1\n")
                            clan2.write("Arquero: 1\n")
                            clan2.write("Gigante: 1\n")
                            clan2.write("Caballero: 1\n")
                            clan2.close()
                            
                        elif clanIngresado==3:
                            clan3=open(rutaClan3,"w+")
                            clan3.write(textoIngresado)
                            clan3.write("Espadachin: 1\n")
                            clan3.write("Arquero: 1\n")
                            clan3.write("Gigante: 1\n")
                            clan3.write("Caballero: 1\n")
                            clan3.close()
                                                    
                        textoIngresado=""
                        estado="iniciarJuego"
                elif event.key == pygame.K_BACKSPACE:
                    textoIngresado = textoIngresado[:-1]
                elif len(textoIngresado)<15:
                    textoIngresado += event.unicode

        #Menu del juego pero ya iniciado, para elegir si jugar, ver el historial o regresar
        elif estado=="iniciarJuego":
            if event.type == pygame.MOUSEBUTTONDOWN:
                if botonJugar.collidepoint(event.pos):
                    estado="jugar"
                if regresar.collidepoint(event.pos):
                    estado="seleccionarPartida"
                if botonPartidas.collidepoint(event.pos):
                    historial()
                    cargarColaHistorial()
                    if colaHistorial.size()==0:    
                        textoError=fuente3.render("No hay partidas",True,rojoIntenso)
                        ventana.blit(textoError,(350,300))
                        pygame.display.update()
                        pygame.time.delay(1000)
                        estado="iniciarJuego"
                    else:
                        
                        estado="historial"
                 
        #El primer caso de jugar, donde se seleccionan guerreros e inicializan funciones y listas como los enemigos y el historial de partidas
        elif estado=="jugar":
            if event.type == pygame.MOUSEBUTTONDOWN:
                if botonVida.collidepoint(event.pos):
                    sortear="Vida"
                    guerrerosQuicksort(guerreros)
                elif botonDanio.collidepoint(event.pos):
                    sortear="Danio"
                    guerrerosQuicksort(guerreros)
                elif botonRegresar.collidepoint(event.pos):
                    estado="iniciarJuego"
                for cuadro,guerrero in cuadroGuerrero:
                    if cuadro.collidepoint(event.pos):
                        
                        if guerrero in guerrerosElegidos:
                            guerrerosElegidos.remove(guerrero)
                        else:
                            if len(guerrerosElegidos) < 3:
                                guerrerosElegidos.append(guerrero)
                            else:
                                mensajeError2 = fuente2.render("Máximo de guerreros seleccionados", True, rojoIntenso)
                                ventana.blit(mensajeError2, (170,215))
                                pygame.display.update()
                                pygame.time.delay(2000)
                if len(guerrerosElegidos) == 3:
                    if botonContinuar.collidepoint(event.pos):
                        equipoRojo=random.sample(range(1,5),3)
                        print(equipoRojo)
                        for i in range(len(equipoRojo)):
                            enemigosRonda.append(guerrerosRojos[equipoRojo[i]-1])
                        for i,guerreros in enumerate(guerrerosElegidos):
                            vidasGuerreros.append(guerreros["Vida"]) 
                        for i,enemigos in enumerate(enemigosRonda):
                            vidasEnemigos.append(enemigos["Vida"])
                            ordenEnemigos.append(enemigos["Nombre"])
                        print(vidasGuerreros)
                        print(vidasEnemigos)
                        colaTurnos1()
                        copiaGuerreros=guerrerosElegidos.copy()
                        resultado=""                       
                        historialExcel=openpyxl.load_workbook(rutaExcel)
                        if clanIngresado==1:
                            hojaClan=historialExcel["HistorialClan1"]
                        elif clanIngresado==2:
                            hojaClan=historialExcel["HistorialClan2"]
                        elif clanIngresado==3:
                            hojaClan=historialExcel["HistorialClan3"]   
                        tamanioHistorial=0   
                        for i in hojaClan['A']:
                            if i is not None:    
                                tamanioHistorial+=1
                            else: 
                                break

                        hojaClan.cell(row=tamanioHistorial+1,column=1,value=tamanioHistorial)
                        hojaClan.cell(row=tamanioHistorial+1,column=2,value=copiaGuerreros[0]["Nombre"])
                        hojaClan.cell(row=tamanioHistorial+1,column=3,value=copiaGuerreros[1]["Nombre"])
                        hojaClan.cell(row=tamanioHistorial+1,column=4,value=copiaGuerreros[2]["Nombre"])    

                        historialExcel.save(rutaExcel)
                        historialExcel.close()

                        estado="combate"
                                         
        #Botones con cada accion segun que enemigo se quiere atacar o si se curan dentro del combate de clanes
        elif estado=="combate":
  
            if not yaAtaco:
                bandera=cola.remove_front()
                cola.add_front(bandera)
                
            if bandera[0]=="Rojo":
                ataqueEvaluarAvidos1(cola)
                yaAtaco=True
            elif bandera[0]=="Azul":
                pass

            if len(vidasEnemigos)==0: 
                resultado="Victoria"
                guerreros=[espada,arco,gigante,caballero]
                colaActualizada=False
                
                estado="resultado"
            elif len(vidasGuerreros)==0:
                resultado="Derrota"
                guerreros=[espada,arco,gigante,caballero]
                colaActualizada=False
                estado="resultado"
       
            if event.type == pygame.MOUSEBUTTONDOWN:
                if ataqueDanio.collidepoint(event.pos):
                    elegirAtaque=True
                elif elegirAtaque:
                    if len(vidasEnemigos)>0:
                        if atacar1.collidepoint(event.pos):
                            restar=cola.remove_front()
                            if len(vidasEnemigos)>=1:
                                vidasEnemigos[0]-=restar[1]["Danio"]
                                vidasEnemigos[0]=max(0,vidasEnemigos[0])    
                                accionA="Acción: "+str(restar[1]["Nombre"])+" Azul atacó a "+str(enemigosRonda[0]['Nombre'])+" Rojo"
                            cola.add_front(restar)
                            actualizarEstadoCombate()
                            cola.add_rear(cola.remove_front())
                            actualizarColaTurnos()             
                            accionGuerreros=fuente1d.render(accionA,True,colorTexto)
                            pygame.display.update() 
                            ventana.blit(accionGuerreros,(405,550))
                            pygame.display.update()    
                            pygame.time.delay(1500)   
                            elegirAtaque=False
                            yaAtaco=False                           
                            estado="combate"
                    if len(vidasEnemigos)>1:
                        if atacar2.collidepoint(event.pos):
                            restar=cola.remove_front()
                            if len(vidasEnemigos)>=2:
                                vidasEnemigos[1]-=restar[1]["Danio"]
                                vidasEnemigos[1]=max(0,vidasEnemigos[1])
                                accionA="Acción: "+str(restar[1]["Nombre"])+" Azul atacó a "+str(enemigosRonda[1]['Nombre'])+" Rojo"
                            cola.add_front(restar)
                            actualizarEstadoCombate()
                            cola.add_rear(cola.remove_front())
                            actualizarColaTurnos()
                            accionGuerreros=fuente1d.render(accionA,True,colorTexto)  
                            pygame.display.update() 
                            ventana.blit(accionGuerreros,(405,550))
                            pygame.display.update()    
                            pygame.time.delay(1500)   
                            elegirAtaque=False
                            yaAtaco=False
                            estado="combate"
                    if len(vidasEnemigos)>2:
                        if atacar3.collidepoint(event.pos):
                            restar=cola.remove_front()
                            if len(vidasEnemigos)>=3:
                                vidasEnemigos[2]-=restar[1]["Danio"]
                                vidasEnemigos[2]=max(0,vidasEnemigos[2])
                                accionA="Acción: "+str(restar[1]["Nombre"])+" Azul atacó a "+str(enemigosRonda[2]['Nombre'])+" Rojo"
                            cola.add_front(restar)
                            actualizarEstadoCombate()
                            cola.add_rear(cola.remove_front())
                            actualizarColaTurnos()
                            accionGuerreros=fuente1d.render(accionA,True,colorTexto)
                            pygame.display.update() 
                            ventana.blit(accionGuerreros,(405,550))
                            pygame.display.update()    
                            pygame.time.delay(1500)   
                            elegirAtaque=False
                            yaAtaco=False
                            estado="combate"
                    
                elif curarVida.collidepoint(event.pos):
                    restar=cola.remove_front()
                    cola.add_front(restar)
                    curar(cola)
                    actualizarColaTurnos()
                    accionA="Acción: "+str(restar[1]["Nombre"])+" Azul se curó"
                    accionGuerreros=fuente1d.render(accionA,True,colorTexto)   
                    if vidaMaxima:
                        pygame.display.update() 
                        ventana.blit(accionGuerreros,(405,550))
                        pygame.display.update()    
                        pygame.time.delay(1500)   
                    colaActualizada=False
                    elegirAtaque=False
                    yaAtaco=False
     
        #Botones para recorrer la cola del historial
        elif estado=="historial":
            if event.type == pygame.MOUSEBUTTONDOWN:
                if botonRegresarH.collidepoint(event.pos):
                    estado="iniciarJuego"
                elif arriba.collidepoint(event.pos):                      
                    if ultimaDireccion=="abajo":
                        cargarColaHistorialAlreves()
                    cargarColaHistorialAlreves()
                    ultimaDireccion="arriba"
                    
                elif abajo.collidepoint(event.pos):
                    if ultimaDireccion=="arriba":
                        cargarColaHistorial()
                    cargarColaHistorial()
                    ultimaDireccion="abajo"
            
        #Boton para regresar de la ventana de creditos
        elif estado=="creditos":
            if event.type == pygame.MOUSEBUTTONDOWN:       
                if botonRegresarMenu.collidepoint(event.pos):
                    estado="menu"

        #Teclado para ingresar la contraseña de administrador
        elif estado=="contraseniaAdmin":
            if event.type == pygame.MOUSEBUTTONDOWN:       
                if regresar4.collidepoint(event.pos):
                    estado="menu"
            
            if event.type == pygame.KEYDOWN:
                if event.key == pygame.K_RETURN:
                    
                    if textoContrasenia==contrasenia:
                        estado="admin"
                    else:
                        error=fuente3.render("Contraseña incorrecta",True,rojoIntenso)
                        ventana.blit(error,(300,450))
                        pygame.display.update()
                        pygame.time.delay(1000)                      
                        textoContrasenia=""
                        estado="contraseniaAdmin"
                elif event.key == pygame.K_BACKSPACE:
                    textoContrasenia = textoContrasenia[:-1]
                elif len(textoContrasenia)<20:
                    textoContrasenia += event.unicode
        
        #Botones de administrador y carga de los archivos luego de actualizar alguna configuracion
        elif estado=="admin":
            if event.type == pygame.MOUSEBUTTONDOWN:       
                if regresarMenu.collidepoint(event.pos):
                    estado="menu"
                
                elif indicarNivel.collidepoint(event.pos):
                    if datoAdmin=="escalado":
                        datoAdmin=""
                        
                    else:
                        datoAdmin="escalado"

                    nivelAdmin=None
                if bajo.collidepoint(event.pos):       
                    escalado="bajo"
                    admi=open(rutaAdmin,"r")
                    linea1=admi.readline()
                    a=admi.readline()
                    admi.close()            
                    admi=open(rutaAdmin,"r")
                    primera = admi.readline()
                    admi.close()                 
                    admi=open(rutaAdmin,"w")
                    admi.write(primera)                     
                    admi.write(escalado+"\n")
                    admi.close()  
                    print(escalado)   
                elif medio.collidepoint(event.pos):  
                                        
                    escalado="normal"
                    admi=open(rutaAdmin,"r")
                    linea1=admi.readline()
                    a=admi.readline()
                    admi.close()                
                    admi=open(rutaAdmin,"r")
                    primera = admi.readline()
                    admi.close()               
                    admi=open(rutaAdmin,"w")
                    admi.write(primera)                     
                    admi.write(escalado+"\n") 
                    admi.close()          

                    print(escalado)             
                elif alto.collidepoint(event.pos):
                    
                    escalado="alto"       
                    admi=open(rutaAdmin,"r")
                    linea1=admi.readline()
                    a=admi.readline()
                    admi.close()       
                    admi=open(rutaAdmin,"r")
                    primera = admi.readline()
                    admi.close()        
                    admi=open(rutaAdmin,"w")
                    admi.write(primera)                     
                    admi.write(escalado+"\n")
                    admi.close()  
                    print(escalado)   
                elif indicarDificultad.collidepoint(event.pos):
                    if datoAdmin=="dificultad":
                        datoAdmin=""
                        
                    else:
                        datoAdmin="dificultad"
                    
                    nivelAdmin=None
                if facil.collidepoint(event.pos):
                    dificultad="facil\n"
                    admi = open(rutaAdmin, "r")
                    lineas = admi.readlines()
                    admi.close()
                    lineas[0] = dificultad
                    admi = open(rutaAdmin, "w")
                    admi.writelines(lineas)
                    admi.close()
                    print(dificultad)
                elif dificil.collidepoint(event.pos):
                    dificultad="dificil\n"
                    admi = open(rutaAdmin, "r")
                    lineas = admi.readlines()
                    admi.close()
                    lineas[0] = dificultad
                    admi = open(rutaAdmin, "w")
                    admi.writelines(lineas)
                    admi.close()
                    print(dificultad)
                elif indicarVida.collidepoint(event.pos):
                    if datoAdmin=="vida":
                        datoAdmin=""
                        nivelAdmin=None
                    else:    
                        datoAdmin="vida"
                        nivelAdmin="Vida"
                elif indicarDanio.collidepoint(event.pos):
                    if datoAdmin=="danio":
                        datoAdmin=""
                        nivelAdmin=None
                    else:
                        datoAdmin="danio"
                        nivelAdmin="Danio"
                
                elif menosVida.collidepoint(event.pos):
                    print("Hice algo")
                    aumentarDatosBase(nivelAdmin)
                    datosPersonajes=obtenerDatos()
                    mostrarDatos(ventana, fuente1c, datosPersonajes, nivelAdmin)
                    print(guerrerosDB)
                elif masVida.collidepoint(event.pos):
                    print("Hice algo")
                    disminuirDatosBase(nivelAdmin)
                    datosPersonajes=obtenerDatos()
                    mostrarDatos(ventana, fuente1c, datosPersonajes, nivelAdmin)
                    print(guerrerosDB)
                elif menosDanio.collidepoint(event.pos):
                    print("Hice algo")
                    aumentarDatosBase(nivelAdmin)
                    datosPersonajes=obtenerDatos()
                    mostrarDatos(ventana, fuente1c, datosPersonajes, nivelAdmin)
                elif masDanio.collidepoint(event.pos):
                    print("Hice algo")
                    disminuirDatosBase(nivelAdmin)
                    datosPersonajes=obtenerDatos()
                    mostrarDatos(ventana, fuente1c, datosPersonajes, nivelAdmin)

    #Switch case donde se va cambiando de pantalla y funciones segun lo seleccionado por el usuario
    if estado=="menu":
        menu()
    elif estado=="seleccionarPartida":
        seleccionarpartida()
    elif estado=="ingresarClan":
        ingresarClan()
    elif estado=="iniciarJuego":
        iniciarJuego(clanIngresado)
    elif estado=="historial":
        mostrarInterfazHistorial()
    elif estado=="jugar":
        jugar()
    elif estado=="combate":
        combate()
    elif estado=="creditos":
        creditos()
    elif estado=="resultado":
        mostrarResultado(resultado)
        pygame.time.delay(3000)
        estado="iniciarJuego"
    elif estado=="contraseniaAdmin":
        contraseñaAdmin()
    elif estado=="admin":
        admin()

    #Actualizar la pantalla luego de cada cambio
    pygame.display.flip()
